#Importamos render, redirect y get_object_or_404 para manejo de vistas (renderizacion, redirigir y buscar objetos)
from django.shortcuts import render, redirect, get_object_or_404
#Importamos messages para mensajes flash
from django.contrib import messages
#Importamos los modelos a usar
from .models import Gestor, Expediente
#Importamos los formularios a usar
from retirementApp.forms import GestorForm, ExpedienteForm, CustomLoginForm,    EditarPerfilForm
#Importamos autenticacion, login y logout
from django.contrib.auth import authenticate, login, logout
#Importamos el modelo User
from django.contrib.auth.models import User, Group
#Importamos login_required para proteger vistas y user_passes_test para permisos(se asegura que sea admin)
from django.contrib.auth.decorators import login_required, user_passes_test
#Importamos Q para manejar comparaciones complejas en las consultas
from django.db.models import Q
from datetime import date
#Importamos Paginator para paginación
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
#Importamos HttpResponse para exportación
from django.http import HttpResponse
#Importamos openpyxl para Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

# Create your views here.

#Verifica si el usuario es admin
def es_admin(user):
    #Verifica que el usuario pertenezca al grupo Administrador
    return user.is_authenticated and (user.is_superuser or user.groups.filter(name='Administrador').exists())
    

#Verifica si es gestor
def es_gestor(user):
    return user.is_authenticated and user.groups.filter(name='Gestor').exists()


def inicio(request):
    return render(request, 'index.html')


#Logica de Autenticacion para LOGIN

#views para LOGIN y REGISTRO
def custom_login(request):
    #Si el usuario ya esta autenticado, redirige al inicio
    if request.user.is_authenticated:
        return redirect('inicio')
    #Si el metodo es POST, procesa el formulario para LOGIN
    if request.method == 'POST':
        form = CustomLoginForm(request, data = request.POST)
        #Verifica que el formulario sea valido
        if form.is_valid():
            #Trae los datos limpios del formulario
            username= form.cleaned_data.get('username')
            password= form.cleaned_data.get('password')
            #Autentica al usuario, comprobando credenciales
            user = authenticate(username=username, password=password)
            
            #Si el usuario es valido, lo loguea
            if user is not None:
                #Inicia sesion
                login(request, user)
                messages.success(request, 'Bienvenido/a de nuevo!') #Mensaje de exito
                #Redirigir segun su rol
                if user.groups.filter(name='Administrador').exists():
                    return redirect('gestores')
                #Si es gestor, redirige a los expedientes existentes
                elif user.groups.filter(name='Gestor').exists():
                    return redirect('expedientes') #Redirige a expedientes para gestores
                else:
                    messages.error(request, 'No cuenta con una cuenta o rol asignado, contacte a su Manager')            
            else: #Si no es valido, arrojara error
                messages.error(request, 'Usuario o contraseña incorrectos')
        else:
            messages.error(request, 'Corrija los datos ingresados')
    else:#Si no es POST, muestra el formulario vacio no autenticado
        form = CustomLoginForm()
    #Renderiza el formulario de login
    return render(request, 'auth/login.html', {'form': form, 'title':'Iniciar Sesión'})


#Views para LOGOUT
def custom_logout(request):
    #Si el request es de un usuario autenticado, obtiene su nombre, sino usa 'Usuario'
    name = request.user.first_name if request.user.is_authenticated and request.user.first_name else 'Usuario'
    logout(request)
    #Mensaje de exito
    messages.success(request, f'¡Nos vemos a la vuelta {name}!')
    return redirect('inicio')  #Redirige a inicio


#Vista protegida para registro de usuarios - SOLO ADMIN
@login_required(login_url='login')
@user_passes_test(es_admin)
def register_user(request):
    #Vista para que Admin cree nuevos usuarios
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '').strip()
        password2 = request.POST.get('password2', '').strip()
        rol = request.POST.get('rol', '').strip()
        
        errores = []
        if not all([username, email, password, password2, rol]):
            errores.append('Todos los campos son obligatorios.')
        
        elif password != password2:
            errores.append('Las contraseñas no coinciden.')

        elif User.objects.filter(username=username).exists():
            errores.append('El nombre de usuario ya existe.')

        elif User.objects.filter(email=email).exists():
            errores.append('El correo electrónico ya esta en uso.')
            
        elif len(password) < 6:
            errores.append('La contraseña debe tener al menos 6 caracteres.')
            
        if errores:
            for error in errores:
                messages.error(request, error)
        else:
            try:
                # Creacion del usuario
                nuevo_usuario = User.objects.create_user(
                    username = username,
                    email = email,
                    password = password,
                )    

                #Asignacion de rol
                if rol  == 'Administrador':
                    grupo = Group.objects.get(name='Administrador')
                else:
                    grupo = Group.objects.get(name='Gestor')
                nuevo_usuario.groups.add(grupo)
                nuevo_usuario.save()
                messages.success(request, f'Usuario {username} creado con exito' )
                return redirect('gestores')
            except Exception as error:
                messages.error(request, f'Error al crear usuario: {error}')
            
    data = {
        'title': 'Registrar Usuario'
    }
    return render(request, 'auth/register.html', data)

#** CRUD GESTORES **#
#We're usign function based views for simplicity

#Vista para crear un gestor
@login_required(login_url='login')
@user_passes_test(es_admin)
def crearGestor(request):
    """Vista unificada - Crear Gestor + Usuario"""
    if request.method == 'POST':
        form = GestorForm(request.POST)
        if form.is_valid():
            try:
                #Usar el método save_gestor (como save_user en register)
                gestor = form.save_gestor()
                messages.success(
                    request, 
                    f'Gestor {gestor.nombre} {gestor.apellido} creado exitosamente'
                )
                return redirect('gestores')
            except Exception as error:
                messages.error(request, f'Error al crear gestor: {error}')
        else:
            messages.error(request, 'Corrija los errores en el formulario')
    else:
        form = GestorForm()
    
    data = {
        'titulo': 'Crear Gestor',
        'form': form
    }
    return render(request, 'gestores/createGestor.html', data)

@login_required(login_url= 'login')
#view para lista todos los gestores con filtro y busqueda
def listaGestores(request):
    #Verificar rol
    if es_gestor(request.user):
        #El gestor solo podra ver info limitada - Su propio perfil con los expedientes asignados
        try:
            gestor_usuario = Gestor.objects.get(
                nombre = request.user.first_name,
                apellido = request.user.last_name
            )
            #Filtro para mostrar su propio perfil
            gestores = Gestor.objects.filter(id=gestor_usuario.id)
            create = False
            titulo = 'Mi perfil'
        except Gestor.DoesNotExist:
            messages.error(request, 'No se encontro el gestor asociado a su cuenta. Contacte a su Manager.')
    #Si es Admin podra ver todos los gestores
    elif es_admin(request.user):
        gestores = Gestor.objects.all()
        create = True
        titulo = 'Gestores'
    #Si no tiene rol asignada indicara error
    else:
        messages.error(request, 'No cuenta con permisos para acceder a esta sección.')
        return redirect('login')
    
    #Filtrado para admins
    query = request.GET.get('query', '').strip()
    if query and create:  # Solo si es admin
        gestores = gestores.filter(
            Q(nombre__icontains=query) | Q(apellido__icontains=query) | 
            Q(email__icontains=query) | Q(rut__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(gestores, 10)  # 10 gestores por página
    page = request.GET.get('page')
    
    try:
        gestores_paginados = paginator.page(page)
    except PageNotAnInteger:
        # Si page no es un entero, mostrar la primera página
        gestores_paginados = paginator.page(1)
    except EmptyPage:
        # Si page está fuera de rango, mostrar la última página
        gestores_paginados = paginator.page(paginator.num_pages)
    
    data = {
        'gestores': gestores_paginados,
        'query': query,
        'create': create,
        'title': titulo,
        'total_gestores': gestores.count()
    }
    return render(request, 'gestores/gestores.html', data)

@login_required(login_url = 'login')
@user_passes_test(es_admin)
#view para editar un gestor
def editarGestor(request, id):
    #buscamos gestor por id
    gestor = get_object_or_404(Gestor, id=id)
    #validamos si es post
    if request.method == 'POST':
        #Procesa formulario con modelo existente
        form = GestorForm(request.POST, instance=gestor)
        #Si es valido, guarda
        if form.is_valid():
            form.save()
            messages.success(request, 'Gestor actualizado exitosamente')
            return redirect('gestores') #Redirije a la lista de gestores
        else: 
            messages.error(request, 'Corrija los errores en el formulario')
    else:
        form = GestorForm(instance=gestor) 
            
    data = {
        'titulo':'Editar Gestor',
        'form': form
    }
    return render(request, 'gestores/createGestor.html', data)

# En views.py - NUEVA VISTA:

@login_required(login_url='login')
def editarPerfil(request, id=None):
    """Vista unificada para editar perfil/gestor según rol"""
    
    # Determinar rol del usuario
    es_admin_user = es_admin(request.user)
    es_gestor_user = es_gestor(request.user)
    
    # Lógica según rol
    if es_admin_user:
        # ADMIN: Puede editar cualquier gestor
        if id:
            gestor = get_object_or_404(Gestor, id=id)
            titulo = f'Editar Gestor - {gestor.nombre} {gestor.apellido}'
            action = 'editar_gestor'
        else:
            messages.error(request, 'ID de gestor requerido')
            return redirect('gestores')
            
    elif es_gestor_user:
        # GESTOR: Solo puede editar su propio perfil
        try:
            gestor = Gestor.objects.get(
                nombre=request.user.first_name,
                apellido=request.user.last_name
            )
            # Ignorar ID para gestores - siempre su perfil
            titulo = 'Editar Mi Perfil'
            action = 'editar_perfil'
        except Gestor.DoesNotExist:
            messages.error(request, 'No se encontró su perfil de gestor')
            return redirect('inicio')
    else:
        messages.error(request, 'No tiene permisos para acceder a esta sección')
        return redirect('login')
    
    # Procesar formulario
    if request.method == 'POST':
        form = EditarPerfilForm(
            request.POST, 
            instance=gestor, 
            es_admin=es_admin_user
        )
        
        if form.is_valid():
            try:
                gestor_actualizado = form.save()
                
                # Actualizar usuario asociado si es necesario
                if es_admin_user:
                    # Admin puede cambiar nombres - actualizar User
                    try:
                        usuario = User.objects.get(
                            first_name=gestor.nombre,
                            last_name=gestor.apellido
                        )
                        usuario.first_name = gestor_actualizado.nombre
                        usuario.last_name = gestor_actualizado.apellido
                        usuario.email = gestor_actualizado.email
                        usuario.save()
                    except User.DoesNotExist:
                        pass  # Usuario no encontrado, continuar
                
                # Mensaje según rol
                if es_admin_user:
                    messages.success(request, f'Gestor "{gestor_actualizado.nombre} {gestor_actualizado.apellido}" actualizado exitosamente')
                    return redirect('gestores')
                else:
                    messages.success(request, 'Su perfil ha sido actualizado exitosamente')
                    return redirect('gestores')  # Su vista de perfil
                    
            except Exception as error:
                messages.error(request, f'Error al actualizar: {error}')
        else:
            # Mostrar errores del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = EditarPerfilForm(instance=gestor, es_admin=es_admin_user)
    
    data = {
        'form': form,
        'gestor': gestor,
        'title': titulo,
        'action': action,
        'es_admin': es_admin_user,
        'puede_cambiar_nombres': es_admin_user,
        'readonly_campos': not es_admin_user
    }
    return render(request, 'gestores/editar_perfil.html', data)


@login_required(login_url='login')
@user_passes_test(es_admin)
#vista para eliminar un gestor
def eliminarGestor(request, id):
    #Buscamos el gestor por id
    gestor = get_object_or_404(Gestor, id=id)
    #Si es post, eliminamos
    if request.method == 'POST':
        try:
            #Guardamos el nombre en una variable para el mensaje
            nombre = f'{gestor.nombre} {gestor.apellido}'
            
            # Buscar y eliminar usuario asociado (silenciosamente)
            username = f'{gestor.nombre.lower()}.{gestor.apellido.lower()}'
            user = User.objects.filter(username=username).first()
            if user:
                user.delete()
            
            # Eliminar el gestor
            gestor.delete()
            messages.success(request, f'Gestor {nombre} ha sido eliminado correctamente')
            return redirect('gestores')
        
        except Exception as error:
            messages.error(request, f'Error al eliminar gestor: {error}')
            return redirect('gestores')
    
    # Si no es POST, redirigir a gestores
    return redirect('gestores')


#* CRUD EXPEDIENTES *#

@login_required(login_url='login')
def listaExpedientes(request):
    #Verifica roles
    if es_gestor(request.user):
        #Gestor vera solo expedientes asignados a el
        try:
            gestor_usuario = Gestor.objects.get(
                nombre=request.user.first_name, 
                apellido= request.user.last_name
            )
            expedientes = Expediente.objects.filter(gestor = gestor_usuario).select_related('gestor')
            create = False
            titulo = 'Mis expedientes asignados'
        except Gestor.DoesNotExist:
            expedientes = Expediente.objects.none()
            create = False
            titulo = 'Expedientes'
            messages.warning(request, 'No se encontro su perfil de gestor')
    #Si es admin podra crear expediente, ver expedientes, editarlos y eliminarlos
    elif es_admin(request.user):
        expedientes = Expediente.objects.all()
        create = True
        titulo = 'Gestion de Expedientes'
    #Si no cuenta con ningun rol arrojara mensaje y sera redirigido a login 
    else:
        messages.error(request, 'No cuenta con permisos para acceder a esta seccion')
        return redirect('login')

    #Filtro para busquedas
    query = request.GET.get('query', '').strip()
    if query:
            expedientes = expedientes.filter(
            Q(titulo__icontains=query) | Q(tipo_pension__icontains=query)
            |Q(gestor__nombre__icontains=query) | Q(gestor__apellido__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(expedientes, 8)  # 8 expedientes por página
    page = request.GET.get('page')
    
    try:
        expedientes_paginados = paginator.page(page)
    except PageNotAnInteger:
        # Si page no es un entero, mostrar la primera página
        expedientes_paginados = paginator.page(1)
    except EmptyPage:
        # Si page está fuera de rango, mostrar la última página
        expedientes_paginados = paginator.page(paginator.num_pages)
    
    #Informacion de fecha actual para comparar vencimientos
    today = date.today()
    
    data = {
        'expedientes': expedientes_paginados,
        'query':query,
        'create':create,
        'title': titulo,
        'today': today,
        'total_expedientes': expedientes.count()
        }
    return render(request, 'expedientes/expedientes.html', data)    



@login_required(login_url='login')
@user_passes_test(es_admin)
def crearExpediente(request):
    """Vista para crear expediente - Solo administradores"""
    
    if request.method == 'POST':
        form = ExpedienteForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                expediente = form.save()
                messages.success(request, f'Expediente "{expediente.titulo}" creado exitosamente')
                return redirect('expedientes')
            except Exception as error:
                messages.error(request, f'Error al crear expediente: {error}')
        else:
            messages.error(request, 'Corrija los errores en el formulario')
    else:
        form = ExpedienteForm()
        
    data = {
        'titulo': 'Crear Expediente',
        'form': form,
        'gestores': Gestor.objects.all()
    }
    return render(request, 'expedientes/form.html', data)



@login_required(login_url='login')
def detalleExpediente(request, id):
    #Obtiene el expediente o muestra 404 
    expediente = get_object_or_404(Expediente, id=id)
    # Verificar permisos
    if es_gestor(request.user):
        # Gestor solo puede ver expedientes asignados a él
        try:
            gestor_obj = Gestor.objects.get(
                nombre=request.user.first_name,
                apellido=request.user.last_name
            )
            if expediente.gestor != gestor_obj:
                messages.error(request, 'No tiene permisos para ver este expediente')
                return redirect('expedientes')
        except Gestor.DoesNotExist:
            messages.error(request, 'No se encontró su perfil de gestor')
            return redirect('expedientes')
    elif not es_admin(request.user):
        messages.error(request, 'No tiene permisos para acceder a esta sección')
        return redirect('login')
    
    # Información adicional para el template
    hoy = date.today()
    #resta los dias de vencimiento menos now y da los dias restantes
    dias_para_vencer = (expediente.fecha_vencimiento - hoy).days 
    
    data = {
        'expediente': expediente,
        'titulo': f'Detalle - {expediente.titulo}',
        'puede_editar': es_admin(request.user),
        'puede_eliminar': es_admin(request.user),
        'today': hoy,
        'dias_para_vencer': dias_para_vencer,
        'vencido': expediente.fecha_vencimiento <= hoy
    }
    return render(request, 'expedientes/detalleExpediente.html', data)


@login_required(login_url='login')
@user_passes_test(es_admin)  # Solo Admin puede editar
def editarExpediente(request, id):
    #El expediente solo podra ser eliminado por el Admin
    expediente = get_object_or_404(Expediente, id=id)
    #Si es post, procesa el formulario
    if request.method == 'POST':
        #Recibe los datos, incluyendo archivos e instancia a expediente
        form = ExpedienteForm(request.POST, request.FILES, instance=expediente)
        if form.is_valid():
            try: #Cuando el formato es valido lo guarda y maneja la respuesta contraria con except
                expediente_actualizado = form.save()
                messages.success(request, f'Expediente "{expediente_actualizado.titulo}" actualizado exitosamente')
                return redirect('expedientes')
            except Exception as error:
                messages.error(request, f'Error al actualizar expediente: {error}')
        else:
            # Mostrar errores específicos del formulario
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ExpedienteForm(instance=expediente)
    
    data = {
        'form': form,
        'expediente': expediente,
        'titulo': f'Editar - {expediente.titulo}',
        'action': 'editar',
        'gestores': Gestor.objects.all()  # Para el select del gestor
    }
    return render(request, 'expedientes/form.html', data)


@login_required(login_url='login')
@user_passes_test(es_admin)  # Solo Admin puede eliminar
def eliminarExpediente(request, id):
    #Se obtiene el expediente o muestra 404
    expediente = get_object_or_404(Expediente, id=id)
    
    if request.method == 'POST':
        try:
            # Guarda datos para el mensaje antes de eliminarlo
            titulo_expediente = expediente.titulo
            #trae el gestor asociado a este expediente
            gestor_nombre = f"{expediente.gestor.nombre} {expediente.gestor.apellido}"
            
            # Eliminar archivo si existe
            if expediente.documentos:
                try:
                    expediente.documentos.delete()
                except:
                    pass  # Si no se puede eliminar el archivo, continuar
            
            expediente.delete()
            messages.success(request, f'Expediente "{titulo_expediente}" de {gestor_nombre} eliminado exitosamente')
        except Exception as error:
            messages.error(request, f'Error al eliminar expediente: {error}')
        
        return redirect('expedientes')
    
    # Si es GET, también redirigir a la lista
    return redirect('expedientes')


# ===== VISTAS DE EXPORTACIÓN =====

@login_required(login_url='login')
@user_passes_test(es_admin)  # Solo Admin puede exportar
def exportar_gestores_excel(request):
    """Exportar gestores a Excel - Solo Administradores"""
    
    if not openpyxl:
        messages.error(request, 'La funcionalidad de exportación no está disponible. Instale openpyxl.')
        return redirect('gestores')
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gestores"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Headers
    headers = ['ID', 'RUT', 'Nombre', 'Apellido', 'Email']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Datos
    gestores = Gestor.objects.all().order_by('nombre', 'apellido')
    for row_num, gestor in enumerate(gestores, 2):
        ws.cell(row=row_num, column=1, value=gestor.id)
        ws.cell(row=row_num, column=2, value=gestor.rut)
        ws.cell(row=row_num, column=3, value=gestor.nombre)
        ws.cell(row=row_num, column=4, value=gestor.apellido)
        ws.cell(row=row_num, column=5, value=gestor.email)
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gestores_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response


@login_required(login_url='login')
@user_passes_test(es_admin)  # Solo Admin puede exportar
def exportar_expedientes_excel(request):
    
    #Revisa que exista la libreria de openpyxl en el project
    if not openpyxl:
        messages.error(request, 'La funcionalidad de exportación no está disponible. Instale openpyxl.')
        return redirect('expedientes')
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expedientes"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_alignment = Alignment(horizontal="center")
    
    # Los headers, seran usados para identificar la primera fila, los titulos de las columnas
    headers = ['ID', 'Título', 'Tipo Pensión', 'Estado', 'Fecha Inicio', 'Fecha Vencimiento', 'Gestor', 'Email Gestor']
    #Hacemos un loop para recorres los headers y asignarlos a la primera fila con enumerate (que se inicializa en 1)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Datos, estos son los datos que se van a exportar a excel, los trae desde Expediente y los ordena con gestor y fecha de inicio
    expedientes = Expediente.objects.select_related('gestor').all().order_by('-fecha_inicio')

    #Loop para recorrer los expedientes y asignar los valores a las celdas correcpondientes (enumerate se inicializa en 2 para saltar la fila de los headers)
    for row_num, expediente in enumerate(expedientes, 2):
        #itera sobre cada expediente y asigna los valors a las celdas que corresponen
        ws.cell(row=row_num, column=1, value=expediente.id)
        ws.cell(row=row_num, column=2, value=expediente.titulo)
        ws.cell(row=row_num, column=3, value=expediente.tipo_pension)
        ws.cell(row=row_num, column=4, value=expediente.get_estado_expediente_display())
        ws.cell(row=row_num, column=5, value=expediente.fecha_inicio.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=6, value=expediente.fecha_vencimiento.strftime('%d/%m/%Y'))
        ws.cell(row=row_num, column=7, value=f'{expediente.gestor.nombre} {expediente.gestor.apellido}')
        ws.cell(row=row_num, column=8, value=expediente.gestor.email)
    
    # Ajustar ancho de columnas, itera sobre las columnas para ajustar el ancho segun el contenido
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        #Loop para iterar y ajustar el ancho de cada celda en la columna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
            # El ancho maximo sera 50 para evitar columnas excesivamente anchas
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="expedientes_{date.today().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response